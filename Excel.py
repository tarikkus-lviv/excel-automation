import openpyxl
from openpyxl.styles import Font
import datetime
import re

filename = 'spendings.xlsx'

wb = openpyxl.load_workbook(filename)
ws_names = wb.sheetnames

print(wb.sheetnames)

sheetname = input('What sheet do you need? T/O: ')
sheet_index = ws_names.index(sheetname)

wb.active = sheet_index

def counting():

    for col in wb.active.iter_cols(min_row=2, max_row=1000, max_col=4, values_only=True):
        for i, cell in enumerate(col):
            row = i + 1

            if cell is not None:
                None

            else:
                all = input('\nWhat did you buy? ')

                if all == ' ':
                    return row
                else:
                    cost = all.split(' ')
                    cost = int(cost[len(cost) - 1])

                    item = all.replace(str(cost), '')

                    wb.active['A'+str(row)] = item
                    wb.active['B'+str(row)] = cost

                Bold = Font(size=13, bold=True)
                wb.active['D1'].font = Bold
                wb.active['D1'] = 'TOTAL'
                wb.active['D2'] = '=SUM(B1:B100)'

                if (wb.active['A'+str(row)]).value != '.':
                    wb.active['C'+str(row)] = datetime.datetime.strptime("2020/08/11, 14:35", "%Y/%m/%d, %H:%M").now()

                else:
                    wb.active['C'+str(row)] == ' '

    return row

counting()

wb.save(filename)

# я дізнаюся рядок , потімпишу, що купив і додаю у клітинку, пишу ціну і додаю в іншу клітинку
#
# тепер мені потрібно створити змінну row у якій змінюватимуться значення із кожним колом.
# і в кожному колі додавати до row 1 і записувати значення, а якщо хзначення == no або ні
# то переривати програму і все зберігати.
